from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.responses import FileResponse, JSONResponse
from starlette.middleware.base import BaseHTTPMiddleware
import subprocess
import tempfile
import os
from docx import Document
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import fitz
from PIL import Image
import csv
import json
import openpyxl
import google.generativeai as genai
import re
from pypdf import PdfWriter, PdfReader

app = FastAPI()

genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))

@app.post("/analyze-spec")
async def analyze_spec(
    filename: str = Form(...),
    target_format: str = Form(...),
    spec: str = Form(...),
    filesize: int = Form(0)
):
    if not spec.strip():
        return JSONResponse({"error": "No spec provided"})

    source_ext = filename.split(".")[-1].lower()
    
    prompt = f"""You are a file conversion expert. A user wants to convert a file and has given these instructions.

File: {filename} ({filesize} bytes)
Converting from: {source_ext} to: {target_format}
User instruction: "{spec}"

Return ONLY a JSON object with these fields (use null for anything that doesn't apply):
{{
  "quality": <0-100 integer or null>,
  "width": <pixels integer or null>,
  "height": <pixels integer or null>,
  "max_size_kb": <integer or null>,
  "dpi": <integer or null>,
  "grayscale": <true/false>,
  "summary": "<one sentence describing what will be done>"
}}

Only return the JSON, nothing else."""

    try:
        model = genai.GenerativeModel("gemini-2.5-flash")
        response = model.generate_content(prompt)
        text = response.text.strip()
        text = re.sub(r'```json|```', '', text).strip()
        settings = json.loads(text)
        return JSONResponse(settings)
    except Exception as e:
        return JSONResponse({"error": f"AI analysis failed: {str(e)}"})
    
class CORSMiddlewareManual(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.method == "OPTIONS":
            response = JSONResponse(content={})
            response.headers["Access-Control-Allow-Origin"] = "*"
            response.headers["Access-Control-Allow-Methods"] = "*"
            response.headers["Access-Control-Allow-Headers"] = "*"
            return response
        response = await call_next(request)
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Methods"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "*"
        return response

app.add_middleware(CORSMiddlewareManual)

COMING_SOON = {'mp4', 'mp3', 'wav', 'gif', 'webm', 'mov', 'avi', 'flac', 'ogg', 'm4a', 'mpeg', 'mpg', 'mkv', 'aac', 'opus', 'wma'}

@app.get("/")
def root():
    return {"message": "ConvertR backend is running"}

@app.get("/models")
def list_models():
    try:
        models = genai.list_models()
        return {"models": [m.name for m in models]}
    except Exception as e:
        return {"error": str(e)}
    
@app.post("/convert")
async def convert(
    file: UploadFile = File(...),
    target_format: str = Form("pdf"),
    spec: str = Form(""),
    ai_quality: str = Form(""),
    ai_width: str = Form(""),
    ai_height: str = Form(""),
    ai_max_size_kb: str = Form(""),
    ai_grayscale: str = Form("false")
):
    source_ext = file.filename.split(".")[-1].lower()
    target_ext = target_format.lower()

    if source_ext in COMING_SOON or target_ext in COMING_SOON:
        return JSONResponse({"error": "Video and audio conversion coming soon — stay tuned!"})

    tmp_dir = tempfile.mkdtemp()
    input_path = os.path.join(tmp_dir, file.filename)
    base_name = file.filename.rsplit(".", 1)[0]
    output_filename = base_name + "." + target_ext
    output_path = os.path.join(tmp_dir, output_filename)

    with open(input_path, "wb") as f:
        f.write(await file.read())

    quality = int(ai_quality) if ai_quality else 85
    width = int(ai_width) if ai_width else None
    height = int(ai_height) if ai_height else None
    max_size_kb = int(ai_max_size_kb) if ai_max_size_kb else None
    grayscale = ai_grayscale.lower() == 'true'

    try:
        if source_ext in {'png', 'jpg', 'jpeg', 'webp', 'bmp', 'tiff'} and target_ext in {'png', 'jpg', 'jpeg', 'webp', 'bmp', 'tiff', 'pdf'}:
            img = Image.open(input_path)
            if grayscale:
                img = img.convert('L')
            elif img.mode in ('RGBA', 'LA', 'P') and target_ext in ('jpg', 'jpeg'):
                img = img.convert('RGB')
            if width and height:
                img = img.resize((width, height), Image.LANCZOS)
            elif width:
                ratio = width / img.width
                img = img.resize((width, int(img.height * ratio)), Image.LANCZOS)
            elif height:
                ratio = height / img.height
                img = img.resize((int(img.width * ratio), height), Image.LANCZOS)
            if target_ext == 'pdf':
                img.save(output_path, 'PDF')
            elif target_ext in ('jpg', 'jpeg'):
                img.save(output_path, 'JPEG', quality=quality)
            elif target_ext == 'webp':
                img.save(output_path, 'WEBP', quality=quality)
            else:
                img.save(output_path)

            if max_size_kb and os.path.exists(output_path):
                current_size = os.path.getsize(output_path) / 1024
                if current_size > max_size_kb and target_ext in ('jpg', 'jpeg', 'webp'):
                    q = quality
                    while current_size > max_size_kb and q > 10:
                        q -= 5
                        img.save(output_path, quality=q)
                        current_size = os.path.getsize(output_path) / 1024

        elif source_ext == 'docx' and target_ext == 'pdf':
            doc = Document(input_path)
            c = canvas.Canvas(output_path, pagesize=letter)
            width_pt, height_pt = letter
            y = height_pt - 50
            for para in doc.paragraphs:
                if y < 50:
                    c.showPage()
                    y = height_pt - 50
                c.drawString(50, y, para.text[:100])
                y -= 20
            c.save()

        elif source_ext == 'pdf' and target_ext == 'txt':
            doc = fitz.open(input_path)
            text = ""
            for page in doc:
                text += page.get_text()
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(text)

        elif source_ext == 'txt' and target_ext == 'pdf':
            with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            c = canvas.Canvas(output_path, pagesize=letter)
            width_pt, height_pt = letter
            y = height_pt - 50
            for line in lines:
                if y < 50:
                    c.showPage()
                    y = height_pt - 50
                c.drawString(50, y, line.strip()[:100])
                y -= 20
            c.save()

        elif source_ext == 'csv' and target_ext == 'json':
            rows = []
            with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            with open(output_path, 'w') as f:
                json.dump(rows, f, indent=2)

        elif source_ext == 'json' and target_ext == 'csv':
            with open(input_path, 'r') as f:
                data = json.load(f)
            if isinstance(data, list) and len(data) > 0:
                with open(output_path, 'w', newline='') as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)

        elif source_ext == 'csv' and target_ext == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(output_path)

        elif source_ext == 'xlsx' and target_ext == 'csv':
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active
            with open(output_path, 'w', newline='') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)

        else:
            return JSONResponse({"error": f"Conversion from {source_ext} to {target_ext} not supported yet"})

        if os.path.exists(output_path):
            return FileResponse(
                output_path,
                filename=output_filename,
                media_type="application/octet-stream"
            )
        else:
            return JSONResponse({"error": "Conversion failed"})

    except Exception as e:
        return JSONResponse({"error": f"Conversion failed: {str(e)}"})
    
@app.post("/merge")
async def merge(
    files: list[UploadFile] = File(...),
    page_order: str = Form(""),
):
    tmp_dir = tempfile.mkdtemp()
    writer = PdfWriter()
    input_paths = []

    for file in files:
        input_path = os.path.join(tmp_dir, file.filename)
        with open(input_path, "wb") as f:
            f.write(await file.read())
        input_paths.append((file.filename, input_path))

    # Parse page order if provided
    # Format: "file1.pdf:0,1,2|file2.pdf:0,1"
    page_map = {}
    if page_order:
        for part in page_order.split("|"):
            if ":" in part:
                fname, pages = part.split(":", 1)
                page_map[fname] = [int(p) for p in pages.split(",")]

    try:
        for filename, input_path in input_paths:
            ext = filename.split(".")[-1].lower()

            # Convert images to PDF first
            if ext in {'png', 'jpg', 'jpeg', 'webp', 'bmp'}:
                img = Image.open(input_path)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                pdf_path = input_path + ".pdf"
                img.save(pdf_path, 'PDF')
                input_path = pdf_path

            reader = PdfReader(input_path)
            pages_to_include = page_map.get(filename, list(range(len(reader.pages))))
            for page_num in pages_to_include:
                if page_num < len(reader.pages):
                    writer.add_page(reader.pages[page_num])

        output_path = os.path.join(tmp_dir, "merged.pdf")
        with open(output_path, "wb") as f:
            writer.write(f)

        return FileResponse(
            output_path,
            filename="merged.pdf",
            media_type="application/octet-stream"
        )

    except Exception as e:
        return JSONResponse({"error": f"Merge failed: {str(e)}"})
    
@app.post("/preview-pages")
async def preview_pages(files: list[UploadFile] = File(...)):
    tmp_dir = tempfile.mkdtemp()
    pages_info = []

    try:
        for file in files:
            input_path = os.path.join(tmp_dir, file.filename)
            with open(input_path, "wb") as f:
                f.write(await file.read())

            ext = file.filename.split(".")[-1].lower()

            if ext in {'png', 'jpg', 'jpeg', 'webp', 'bmp'}:
                pages_info.append({
                    "filename": file.filename,
                    "page_num": 0,
                    "total_pages": 1
                })
            elif ext == 'pdf':
                doc = fitz.open(input_path)
                num_pages = len(doc)
                for i in range(num_pages):
                    pages_info.append({
                        "filename": file.filename,
                        "page_num": i,
                        "total_pages": num_pages
                    })
                doc.close()

        return JSONResponse({"pages": pages_info})
    except Exception as e:
        return JSONResponse({"error": str(e)})


@app.post("/page-thumbnail")
async def page_thumbnail(file: UploadFile = File(...), page_num: int = Form(0)):
    tmp_dir = tempfile.mkdtemp()
    input_path = os.path.join(tmp_dir, file.filename)
    with open(input_path, "wb") as f:
        f.write(await file.read())

    ext = file.filename.split(".")[-1].lower()
    thumb_path = os.path.join(tmp_dir, "thumb.png")

    try:
        if ext in {'png', 'jpg', 'jpeg', 'webp', 'bmp'}:
            img = Image.open(input_path)
            img.thumbnail((300, 300))
            img.save(thumb_path, "PNG")
        elif ext == 'pdf':
            doc = fitz.open(input_path)
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(0.5, 0.5))
            pix.save(thumb_path)
            doc.close()

        return FileResponse(thumb_path, media_type="image/png")
    except Exception as e:
        return JSONResponse({"error": str(e)})